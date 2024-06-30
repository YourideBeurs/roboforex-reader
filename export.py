from openpyxl import Workbook
from tqdm import tqdm

def export_to_xlsx(filename: str, data: list):
    
    workbook = Workbook()
    sheet = workbook.active

    sheet.cell(1, 1).value = 'Date'
    sheet.cell(1, 2).value = 'Balance'
    sheet.cell(1, 3).value = 'Deposit/Withdrawal'

    row = 2

    for i in tqdm(range(0, len(data))):
        sheet.cell(i + row, 1).value = data[i][0]
        sheet.cell(i + row, 2).value = data[i][1]
        sheet.cell(i + row, 3).value = data[i][2]

    workbook.save(filename=filename)